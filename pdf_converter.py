#!/usr/bin/env python3
"""
PDF to Excel Converter
Convert PDF documents to formatted Excel spreadsheets.
"""

import sys
import logging
import argparse
from pathlib import Path
from typing import Optional, List, Dict, Any

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Try to import tkinter for file dialogs
try:
    import tkinter as tk
    from tkinter import filedialog
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class PDFtoExcelConverter:
    """Convert PDF files to Excel spreadsheets."""

    def __init__(self, input_path: str, output_path: str):
        """
        Initialize converter.
        
        Args:
            input_path: Path to input PDF file
            output_path: Path to output Excel file
        """
        self.input_path = Path(input_path)
        self.output_path = Path(output_path)
        self.tables: List[pd.DataFrame] = []
        self.metadata: Dict[str, Any] = {}

    def extract_tables(self) -> bool:
        """
        Extract tables from PDF.
        
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Opening PDF: {self.input_path}")
            
            with pdfplumber.open(self.input_path) as pdf:
                self.metadata['total_pages'] = len(pdf.pages)
                logger.info(f"Total pages: {len(pdf.pages)}")
                
                for page_num, page in enumerate(pdf.pages, 1):
                    tables = page.extract_tables()
                    
                    if tables:
                        logger.info(f"Found {len(tables)} table(s) on page {page_num}")
                        
                        for table in tables:
                            df = pd.DataFrame(table[1:], columns=table[0])
                            self.tables.append(df)
                    else:
                        logger.warning(f"No tables found on page {page_num}")
            
            if not self.tables:
                logger.warning("No tables extracted from PDF")
                return False
            
            logger.info(f"Successfully extracted {len(self.tables)} table(s)")
            return True
            
        except Exception as e:
            logger.error(f"Error extracting tables: {str(e)}")
            return False

    def create_excel(self) -> bool:
        """
        Create formatted Excel file from extracted tables.
        
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Creating Excel file: {self.output_path}")
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add metadata sheet
            self._add_metadata_sheet(wb)
            
            # Add table sheets
            for idx, df in enumerate(self.tables, 1):
                sheet_name = f"Table_{idx}"
                ws = wb.create_sheet(sheet_name)
                self._format_sheet(ws, df)
            
            wb.save(self.output_path)
            logger.info(f"Excel file created successfully: {self.output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            return False

    def _add_metadata_sheet(self, wb: Workbook) -> None:
        """
        Add metadata sheet to workbook.
        
        Args:
            wb: Openpyxl Workbook object
        """
        ws = wb.create_sheet("Metadata", 0)
        ws['A1'] = "PDF to Excel Conversion Metadata"
        ws['A1'].font = Font(bold=True, size=12)
        
        metadata_rows = [
            ['Source File', str(self.input_path)],
            ['Output File', str(self.output_path)],
            ['Total Pages', self.metadata.get('total_pages', 'N/A')],
            ['Tables Extracted', len(self.tables)],
        ]
        
        for row_idx, (key, value) in enumerate(metadata_rows, 3):
            ws[f'A{row_idx}'] = key
            ws[f'B{row_idx}'] = value
            ws[f'A{row_idx}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _format_sheet(self, ws, df: pd.DataFrame) -> None:
        """
        Format worksheet with dataframe and styling.
        
        Args:
            ws: Openpyxl worksheet object
            df: Pandas DataFrame to add to sheet
        """
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def convert(self) -> bool:
        """
        Execute full conversion process.
        
        Returns:
            True if successful, False otherwise
        """
        logger.info("Starting PDF to Excel conversion...")
        
        if not self.extract_tables():
            return False
        
        if not self.create_excel():
            return False
        
        logger.info("Conversion completed successfully!")
        return True


def select_file_dialog(title: str = "Select PDF File", file_type: str = "PDF files") -> Optional[str]:
    """
    Open file selection dialog.
    
    Args:
        title: Dialog title
        file_type: File type filter
        
    Returns:
        Selected file path or None
    """
    if not TKINTER_AVAILABLE:
        logger.warning("Tkinter not available. Please provide file path manually.")
        return None
    
    try:
        root = tk.Tk()
        root.withdraw()  # Hide the window
        
        if "PDF" in file_type:
            file_path = filedialog.askopenfilename(
                title=title,
                filetypes=[(file_type, "*.pdf"), ("All Files", "*.*")]
            )
        else:
            file_path = filedialog.asksaveasfilename(
                title=title,
                filetypes=[(file_type, "*.xlsx"), ("All Files", "*.*")],
                defaultextension=".xlsx"
            )
        
        root.destroy()
        return file_path if file_path else None
        
    except Exception as e:
        logger.error(f"Error opening file dialog: {str(e)}")
        return None


def select_folder_dialog(title: str = "Select Output Folder") -> Optional[str]:
    """
    Open folder selection dialog.
    
    Args:
        title: Dialog title
        
    Returns:
        Selected folder path or None
    """
    if not TKINTER_AVAILABLE:
        logger.warning("Tkinter not available. Please provide folder path manually.")
        return None
    
    try:
        root = tk.Tk()
        root.withdraw()  # Hide the window
        
        folder_path = filedialog.askdirectory(title=title)
        root.destroy()
        
        return folder_path if folder_path else None
        
    except Exception as e:
        logger.error(f"Error opening folder dialog: {str(e)}")
        return None


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Convert PDF files to Excel spreadsheets'
    )
    parser.add_argument(
        '--input', '-i',
        help='Path to input PDF file'
    )
    parser.add_argument(
        '--output', '-o',
        help='Path to output Excel file'
    )
    parser.add_argument(
        '--interactive', '-int',
        action='store_true',
        help='Open file/folder selection dialogs'
    )
    
    args = parser.parse_args()
    
    # Use interactive mode if flag is set or no arguments provided
    if args.interactive or (not args.input and not args.output):
        print("\n" + "="*60)
        print("PDF to Excel Converter - Interactive Mode")
        print("="*60 + "\n")
        
        if not TKINTER_AVAILABLE:
            print("❌ Tkinter not available on this system.")
            print("Please install tkinter or provide file paths manually:")
            print(f"  python pdf_converter.py --input input.pdf --output output.xlsx\n")
            sys.exit(1)
        
        # Select input PDF
        print("📁 Step 1: Select PDF file to convert")
        input_path = select_file_dialog("Select PDF File", "PDF files")
        
        if not input_path:
            print("❌ No PDF file selected. Exiting.")
            sys.exit(1)
        
        print(f"✓ Selected: {input_path}\n")
        
        # Select output folder
        print("📁 Step 2: Select where to save Excel file")
        output_folder = select_folder_dialog("Select Output Folder")
        
        if not output_folder:
            output_folder = str(Path(input_path).parent)
            print(f"✓ Using PDF folder: {output_folder}\n")
        else:
            print(f"✓ Selected: {output_folder}\n")
        
        # Generate output filename
        input_name = Path(input_path).stem
        output_path = str(Path(output_folder) / f"{input_name}_converted.xlsx")
        
        args.input = input_path
        args.output = output_path
    
    # Validate inputs
    if not args.input or not args.output:
        parser.print_help()
        sys.exit(1)
    
    input_file = args.input
    output_file = args.output
    
    # Validate input file
    if not Path(input_file).exists():
        logger.error(f"Input file not found: {input_file}")
        sys.exit(1)
    
    # Create converter and execute
    print("\n" + "="*60)
    print("Converting PDF to Excel...")
    print("="*60 + "\n")
    
    converter = PDFtoExcelConverter(input_file, output_file)
    success = converter.convert()
    
    if success:
        print("\n" + "="*60)
        print("✓ Conversion completed successfully!")
        print("="*60)
        print(f"\nOutput file: {output_file}\n")
    else:
        print("\n❌ Conversion failed!")
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
