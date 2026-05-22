#!/usr/bin/env python3
"""
Create sample Excel file based on the bank register PDF structure.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime


def create_sample_excel():
    """
    Create sample Excel file with bank register data.
    """
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add metadata sheet
    ws_meta = wb.create_sheet("Metadata", 0)
    ws_meta['A1'] = "Bank Register - Conversion Metadata"
    ws_meta['A1'].font = title_font
    
    metadata = [
        ['Bank Name', 'Indian Overseas Bank, Sivaganga'],
        ['Service Outlet', 'IOB84 - SIVAGANGAI'],
        ['Account Number', 'IOB4020000065GS/INR'],
        ['Organization', 'Tamil Nadu Pollution Control Board'],
        ['Report Period', '01-04-2026 to 30-04-2026'],
        ['Generated Date', datetime.now().strftime('%d-%m-%Y %H:%M:%S')],
        ['Document Type', 'Bank Statement Register'],
    ]
    
    for row_idx, (key, value) in enumerate(metadata, 3):
        ws_meta[f'A{row_idx}'] = key
        ws_meta[f'B{row_idx}'] = value
        ws_meta[f'A{row_idx}'].font = Font(bold=True)
    
    ws_meta.column_dimensions['A'].width = 25
    ws_meta.column_dimensions['B'].width = 40
    
    # Add transactions sheet
    ws_trans = wb.create_sheet("Transactions", 1)
    
    # Headers
    headers = ['Date', 'Transaction ID', 'Reference Number', 'Particulars', 
               'Debit Amount', 'Credit Amount', 'Balance Amount', 'Currency']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_trans.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Sample data
    data = [
        ['07-04-2026', 'IB170290', 'DD Payment 143637872', 'Account Opening balance', '', '1,23,185.00', '1,23,185.00', 'INR'],
        ['07-04-2026', 'IB170290', 'DD Payment 143637872', 'Brought Forward', '6,960.00', '', '1,30,145.00', 'INR'],
        ['07-04-2026', 'IB1908', 'BY CLG: 548972-0084-07-04', 'Credit Entry', '26,250.00', '', '1,56,395.00', 'INR'],
        ['07-04-2026', 'IB1908', 'BY CLG: 67D-0084-09-04', 'Credit Entry', '15,730.00', '', '1,72,125.00', 'INR'],
        ['13-04-2026', 'IB1908', 'BY CLG: 494020-0084-13-04', 'Debit Entry', '1,19,800.00', '', '2,91,925.00', 'INR'],
        ['13-04-2026', 'IB1908', 'BY CLG: 517443-0084-13-04', 'Debit Entry', '16,950.00', '', '3,08,875.00', 'INR'],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_trans.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Right align numbers
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Set column widths
    widths = [12, 15, 20, 25, 15, 15, 15, 10]
    for col_idx, width in enumerate(widths, 1):
        ws_trans.column_dimensions[chr(64 + col_idx)].width = width
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary", 2)
    ws_summary['A1'] = "Transaction Summary"
    ws_summary['A1'].font = title_font
    
    summary_data = [
        ['Total Transactions', 6],
        ['Total Debit Amount', 1639800],
        ['Total Credit Amount', 1230450],
        ['Opening Balance', 123185],
        ['Closing Balance', 308875],
        ['Net Change', 185690],
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 3):
        ws_summary[f'A{row_idx}'] = label
        ws_summary[f'B{row_idx}'] = value
        ws_summary[f'A{row_idx}'].font = Font(bold=True)
        ws_summary[f'A{row_idx}'].border = border
        ws_summary[f'B{row_idx}'].border = border
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # Save file
    output_file = 'sample_bank_register.xlsx'
    wb.save(output_file)
    print(f"✓ Sample Excel file created: {output_file}")
    print(f"  - Metadata sheet with document information")
    print(f"  - Transactions sheet with bank entries")
    print(f"  - Summary sheet with transaction statistics")


if __name__ == '__main__':
    create_sample_excel()
